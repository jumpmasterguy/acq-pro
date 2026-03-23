"""
FILE 1: ppbe-cycle-one-pager.xlsx
PPBE Cycle Quick Reference — beautiful one-page landscape cheat sheet
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PPBE Cycle Quick Reference"

# ── Brand colors ──
TEAL = "01696F"
NAVY = "0D1B2A"
WHITE = "FFFFFF"
LIGHT_TEAL = "E0F2F3"
LIGHT_NAVY = "E8EAF0"
MED_TEAL = "4DA8AE"
DARK_GRAY = "333333"
MID_GRAY = "666666"
LIGHT_GRAY = "F5F5F5"
BORDER_GRAY = "C0C0C0"
VERY_LIGHT_TEAL = "F0FAFA"

# ── Fills ──
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_teal = PatternFill("solid", fgColor=TEAL)
fill_light_teal = PatternFill("solid", fgColor=LIGHT_TEAL)
fill_light_navy = PatternFill("solid", fgColor=LIGHT_NAVY)
fill_light_gray = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_very_light_teal = PatternFill("solid", fgColor=VERY_LIGHT_TEAL)

# ── Fonts ──
font_title = Font(name="Calibri", size=18, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=10, color="B0C4DE")
font_phase_header = Font(name="Calibri", size=13, bold=True, color=WHITE)
font_label = Font(name="Calibri", size=9, bold=True, color=TEAL)
font_value = Font(name="Calibri", size=9, color=DARK_GRAY)
font_bullet = Font(name="Calibri", size=8, color=MID_GRAY)
font_section_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_table_header = Font(name="Calibri", size=8, bold=True, color=WHITE)
font_table_cell = Font(name="Calibri", size=8, color=DARK_GRAY)
font_term_name = Font(name="Calibri", size=8, bold=True, color=TEAL)
font_term_def = Font(name="Calibri", size=8, color=MID_GRAY)
font_footer = Font(name="Calibri", size=7, italic=True, color=MID_GRAY)

# ── Borders ──
thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
bottom_border = Border(bottom=Side(style="thin", color=BORDER_GRAY))

# ── Alignments ──
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Page Setup (Landscape, fit to one page) ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3
ws.print_options.horizontalCentered = True

# ── Column widths (16 columns: A-P) ──
# Each phase gets 4 columns (B-E, F-I, J-M, N-Q)
# A = left margin
col_widths = {
    'A': 1.5,
    'B': 9, 'C': 9, 'D': 9, 'E': 9,
    'F': 9, 'G': 9, 'H': 9, 'I': 9,
    'J': 9, 'K': 9, 'L': 9, 'M': 9,
    'N': 9, 'O': 9, 'P': 9, 'Q': 9,
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ═══════════════════════════════════════════════════
# ROW 1-2: TITLE BAR (Navy background)
# ═══════════════════════════════════════════════════
for row in [1, 2]:
    for col in range(1, 18):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill_navy

ws.merge_cells("B1:L1")
title_cell = ws["B1"]
title_cell.value = "PPBE Cycle Quick Reference"
title_cell.font = font_title
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("B2:L2")
subtitle_cell = ws["B2"]
subtitle_cell.value = "Planning, Programming, Budgeting & Execution — Defense Acquisitions Academy"
subtitle_cell.font = font_subtitle
subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

# Brand tag on right
ws.merge_cells("M1:Q1")
brand = ws["M1"]
brand.value = "ACQLERATE"
brand.font = Font(name="Calibri", size=14, bold=True, color=MED_TEAL)
brand.alignment = Alignment(horizontal="right", vertical="center")

ws.merge_cells("M2:Q2")
brand2 = ws["M2"]
brand2.value = "Defense Acquisitions Academy"
brand2.font = Font(name="Calibri", size=8, color="B0C4DE")
brand2.alignment = Alignment(horizontal="right", vertical="center")

# ═══════════════════════════════════════════════════
# ROW 3: Spacer
# ═══════════════════════════════════════════════════
ws.row_dimensions[3].height = 6

# ═══════════════════════════════════════════════════
# ROW 4-20: FOUR PHASE BLOCKS
# ═══════════════════════════════════════════════════
# Phase columns: B-E (Planning), F-I (Programming), J-M (Budgeting), N-Q (Execution)
phases = [
    {
        "name": "PLANNING",
        "cols": (2, 5),  # B-E
        "timeframe": "Year N-3 to N-2  |  Jan – Aug",
        "owner": "OSD Policy / Under Secretary of Defense (Policy)",
        "output": "Defense Planning Guidance (DPG)",
        "activities": [
            "Assess global threats & strategic environment",
            "Develop National Defense Strategy guidance",
            "Issue fiscal guidance & planning assumptions",
            "Set force structure & capability priorities",
        ],
        "document": "DPG — Translates strategy into fiscal/force planning guidance for Services",
    },
    {
        "name": "PROGRAMMING",
        "cols": (6, 9),  # F-I
        "timeframe": "Year N-2  |  Feb – Sep",
        "owner": "Military Services / Service Secretariats",
        "output": "Program Objective Memorandum (POM)",
        "activities": [
            "Services build 5-year resource proposals",
            "Allocate $ across programs within fiscal limits",
            "Program Review Groups evaluate alternatives",
            "Issue Program Decision Memoranda (PDMs)",
        ],
        "document": "POM — Each Service's 5-year spending plan, submitted to OSD for review",
    },
    {
        "name": "BUDGETING",
        "cols": (10, 13),  # J-M
        "timeframe": "Year N-1  |  Sep – Feb",
        "owner": "OUSD(C) / OMB / Congress",
        "output": "Budget Estimate Submission (BES)",
        "activities": [
            "Convert approved programs into budget format",
            "OSD/OMB review & issue Program Budget Decisions",
            "Prepare President's Budget (PB) submission",
            "Congressional hearings, markups & appropriation",
        ],
        "document": "BES → President's Budget — Formal budget request to Congress, line-item detail",
    },
    {
        "name": "EXECUTION",
        "cols": (14, 17),  # N-Q
        "timeframe": "Year N  |  Oct 1 – Sep 30",
        "owner": "Program Offices / Comptrollers / DFAS",
        "output": "Obligations & Expenditures",
        "activities": [
            "Apportion & allocate funds to commands/programs",
            "Obligate funds via contracts & orders",
            "Monitor execution rates & under/over-execution",
            "Prepare DD-1414 & SF-133 execution reports",
        ],
        "document": "DD-1414 / SF-133 — Execution reporting to OMB; tracks obligations vs. plan",
    },
]

row = 4  # Start row for phases

for phase in phases:
    c1, c2 = phase["cols"]
    
    # Phase header row (teal)
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    header_cell = ws.cell(row=row, column=c1)
    header_cell.value = phase["name"]
    header_cell.font = font_phase_header
    header_cell.fill = fill_teal
    header_cell.alignment = align_center
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill_teal
        ws.cell(row=row, column=c).border = Border(
            left=Side(style="thin", color=WHITE) if c == c1 else Side(style=None),
            right=Side(style="thin", color=WHITE) if c == c2 else Side(style=None),
        )
    ws.row_dimensions[row].height = 22

    # Timeframe
    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = phase["timeframe"]
    cell.font = Font(name="Calibri", size=8, italic=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor="1A8A90")
    cell.alignment = align_center
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="1A8A90")
    ws.row_dimensions[row].height = 16

    # Light background for content area
    content_start = row + 1
    content_rows = 12  # rows for content
    for r in range(content_start, content_start + content_rows):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill_very_light_teal
            ws.cell(row=r, column=c).border = Border(
                left=Side(style="thin", color=BORDER_GRAY) if c == c1 else Side(style=None),
                right=Side(style="thin", color=BORDER_GRAY) if c == c2 else Side(style=None),
                bottom=Side(style="thin", color=BORDER_GRAY) if r == content_start + content_rows - 1 else Side(style=None),
            )

    # Key Owner
    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = "KEY OWNER"
    cell.font = font_label
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 14

    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = phase["owner"]
    cell.font = font_value
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 22

    # Key Output
    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = "KEY OUTPUT"
    cell.font = font_label
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 14

    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = phase["output"]
    cell.font = Font(name="Calibri", size=9, bold=True, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16

    # Key Activities
    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = "KEY ACTIVITIES"
    cell.font = font_label
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 14

    for activity in phase["activities"]:
        row += 1
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = f"  ▸ {activity}"
        cell.font = font_bullet
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 14

    # Key Document
    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = "KEY DOCUMENT"
    cell.font = font_label
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 14

    row += 1
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = phase["document"]
    cell.font = Font(name="Calibri", size=7.5, italic=True, color=MID_GRAY)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 28

    # Reset row to start (all phases use same rows)
    row = 4  # Reset for next phase

# ═══════════════════════════════════════════════════
# ROWS 18-19: Arrow connectors between phases
# ═══════════════════════════════════════════════════
# (Visual separator row)
arrow_row = 18
ws.row_dimensions[arrow_row].height = 10

# ═══════════════════════════════════════════════════
# ROW 19-25: KEY PPBE DATES TIMELINE
# ═══════════════════════════════════════════════════
timeline_start = 19

# Section header
ws.merge_cells(f"B{timeline_start}:Q{timeline_start}")
cell = ws.cell(row=timeline_start, column=2)
cell.value = "KEY PPBE MILESTONE DATES (Typical Cycle)"
cell.font = font_section_header
cell.fill = fill_navy
cell.alignment = align_center
ws.row_dimensions[timeline_start].height = 20

# Timeline table headers
tl_header = timeline_start + 1
headers = ["Month", "Planning", "Programming", "Budgeting", "Execution", "Key Event"]
header_cols = [
    (2, 3),   # Month
    (4, 6),   # Planning
    (7, 9),   # Programming
    (10, 12), # Budgeting
    (13, 15), # Execution
    (16, 17), # Key Event
]
for i, (h_text, (c1, c2)) in enumerate(zip(headers, header_cols)):
    ws.merge_cells(start_row=tl_header, start_column=c1, end_row=tl_header, end_column=c2)
    cell = ws.cell(row=tl_header, column=c1)
    cell.value = h_text
    cell.font = font_table_header
    cell.fill = fill_teal
    cell.alignment = align_center
    for c in range(c1, c2 + 1):
        ws.cell(row=tl_header, column=c).fill = fill_teal
        ws.cell(row=tl_header, column=c).border = thin_border
ws.row_dimensions[tl_header].height = 16

# Timeline data
timeline_data = [
    ("Jan-Feb", "DPG development", "POM prep begins", "", "", "Strategic guidance issued"),
    ("Mar-May", "DPG finalized", "Services build POM", "", "", "Fiscal guidance released"),
    ("Jun-Aug", "", "POM submitted to OSD", "OSD review begins", "", "Program Review Groups meet"),
    ("Sep-Nov", "", "PDMs issued", "BES prepared", "FY starts Oct 1", "PBDs issued; execution begins"),
    ("Dec-Feb", "", "", "President's Budget to Congress", "Execution continues", "Congressional hearings begin"),
    ("Mar-Sep", "", "", "Appropriations enacted", "Obligation & expenditure", "CRs possible if no approps"),
]

for i, (month, planning, programming, budgeting, execution, event) in enumerate(timeline_data):
    r = tl_header + 1 + i
    row_data = [month, planning, programming, budgeting, execution, event]
    alt_fill = fill_light_teal if i % 2 == 0 else fill_white
    for j, (text, (c1, c2)) in enumerate(zip(row_data, header_cols)):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1)
        cell.value = text
        cell.font = font_table_cell if j > 0 else Font(name="Calibri", size=8, bold=True, color=DARK_GRAY)
        cell.alignment = Alignment(horizontal="center" if j == 0 else "left", vertical="center", wrap_text=True, indent=0 if j == 0 else 1)
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = alt_fill
            ws.cell(row=r, column=c).border = thin_border
    ws.row_dimensions[r].height = 18

# ═══════════════════════════════════════════════════
# ROW 28-34: KEY TERMS
# ═══════════════════════════════════════════════════
terms_start = tl_header + 1 + len(timeline_data) + 1  # row 28

ws.merge_cells(f"B{terms_start}:Q{terms_start}")
cell = ws.cell(row=terms_start, column=2)
cell.value = "KEY TERMS & DEFINITIONS"
cell.font = font_section_header
cell.fill = fill_navy
cell.alignment = align_center
ws.row_dimensions[terms_start].height = 20

terms = [
    ("POM", "Program Objective Memorandum — Each Service's 5-year resource allocation plan, built within fiscal guidance"),
    ("DPG", "Defense Planning Guidance — OSD document translating national strategy into planning/programming guidance for DoD components"),
    ("BES", "Budget Estimate Submission — Detailed budget request from each Service/agency, formatted for OMB and Congressional review"),
    ("FYDP", "Future Years Defense Program — The official 5-year (current + 4) DoD spending plan organized by program elements"),
    ("CR", "Continuing Resolution — Temporary funding authority when Congress fails to pass appropriations; typically at prior-year levels"),
    ("PDM", "Program Decision Memorandum — OSD decisions on Service POM submissions; directs changes before budget formulation"),
    ("PBD", "Program Budget Decision — OSD adjustments during the budget review phase, refining numbers before the President's Budget"),
]

# Two-column layout for terms
for i, (acronym, definition) in enumerate(terms):
    r = terms_start + 1 + (i // 2)
    if i % 2 == 0:
        # Left side
        c1, c2 = 2, 2
        c3, c4 = 3, 9
    else:
        # Right side
        c1, c2 = 10, 10
        c3, c4 = 11, 17

    cell = ws.cell(row=r, column=c1)
    cell.value = acronym
    cell.font = font_term_name
    cell.alignment = Alignment(horizontal="right", vertical="top", indent=0)

    ws.merge_cells(start_row=r, start_column=c3, end_row=r, end_column=c4)
    cell = ws.cell(row=r, column=c3)
    cell.value = definition
    cell.font = font_term_def
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.row_dimensions[r].height = 26

# ═══════════════════════════════════════════════════
# FOOTER ROW
# ═══════════════════════════════════════════════════
footer_row = terms_start + 1 + ((len(terms) + 1) // 2) + 1
ws.merge_cells(f"B{footer_row}:Q{footer_row}")
cell = ws.cell(row=footer_row, column=2)
cell.value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com  |  For educational use — not official DoD guidance"
cell.font = font_footer
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[footer_row].height = 16

# ── Set print area ──
ws.print_area = f"A1:Q{footer_row}"

# ── Gridlines off for print ──
ws.sheet_view.showGridLines = False

# Save
output = "/home/user/workspace/acq-pro/products/pack3-finance-cheat-sheets/ppbe-cycle-one-pager.xlsx"
wb.save(output)
print(f"Saved: {output}")
