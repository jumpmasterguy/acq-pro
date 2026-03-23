"""
FILE 2: color-of-money-decision-tree.xlsx
Color of Money decision tree / reference guide — 2 sheets
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

wb = openpyxl.Workbook()

# ── Brand colors ──
TEAL = "01696F"
NAVY = "0D1B2A"
WHITE = "FFFFFF"
LIGHT_TEAL = "E0F2F3"
VERY_LIGHT_TEAL = "F0FAFA"
DARK_GRAY = "333333"
MID_GRAY = "666666"
BORDER_GRAY = "C0C0C0"
LIGHT_GRAY = "F5F5F5"

# Appropriation colors
COLOR_OM = "2980B9"       # Blue
COLOR_PROC = "27AE60"     # Green
COLOR_RDTE = "8E44AD"     # Purple
COLOR_MILPERS = "D4850A"  # Orange
COLOR_MILCON = "C0392B"   # Red
COLOR_WCF = "7F8C8D"      # Gray
COLOR_REVFUND = "2C3E50"  # Dark

LIGHT_OM = "D6EAF8"
LIGHT_PROC = "D5F5E3"
LIGHT_RDTE = "E8DAEF"
LIGHT_MILPERS = "FDEBD0"
LIGHT_MILCON = "FADBD8"
LIGHT_WCF = "E5E8E8"
LIGHT_REVFUND = "D5D8DC"

# ── Fills ──
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_teal = PatternFill("solid", fgColor=TEAL)
fill_light_teal = PatternFill("solid", fgColor=LIGHT_TEAL)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_light_gray = PatternFill("solid", fgColor=LIGHT_GRAY)

# ── Fonts ──
font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=9, color="B0C4DE")
font_brand = Font(name="Calibri", size=12, bold=True, color="4DA8AE")
font_section = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_header = Font(name="Calibri", size=9, bold=True, color=WHITE)
font_body = Font(name="Calibri", size=9, color=DARK_GRAY)
font_body_bold = Font(name="Calibri", size=9, bold=True, color=DARK_GRAY)
font_big_q = Font(name="Calibri", size=14, bold=True, color=TEAL)
font_branch = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_arrow = Font(name="Calibri", size=12, bold=True, color=TEAL)
font_sub_label = Font(name="Calibri", size=8, bold=True, color=TEAL)
font_sub_body = Font(name="Calibri", size=8, color=MID_GRAY)
font_footer = Font(name="Calibri", size=7, italic=True, color=MID_GRAY)
font_note = Font(name="Calibri", size=8, italic=True, color=MID_GRAY)
font_red_bold = Font(name="Calibri", size=8, bold=True, color="C0392B")

# ── Borders ──
thin = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

# ── Alignments ──
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_indent = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
al_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

def apply_fill(ws, row, c1, c2, fill):
    for c in range(c1, c2+1):
        ws.cell(row=row, column=c).fill = fill

def section_hdr(ws, row, text, c1=2, c2=10):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = text
    cell.font = font_section
    cell.fill = fill_navy
    cell.alignment = ac
    apply_fill(ws, row, c1, c2, fill_navy)
    ws.row_dimensions[row].height = 22
    return row + 1

# ═══════════════════════════════════════════════════
# SHEET 1: DECISION TREE
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Decision Tree"
ws1.sheet_view.showGridLines = False

ws1.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws1.page_setup.orientation = "landscape"
ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
ws1.page_setup.fitToHeight = 1
ws1.page_setup.fitToWidth = 1
ws1.page_margins.left = 0.3
ws1.page_margins.right = 0.3
ws1.page_margins.top = 0.3
ws1.page_margins.bottom = 0.3

# Column widths
col_w = {'A': 1.5, 'B': 5, 'C': 18, 'D': 14, 'E': 12, 'F': 12, 'G': 24, 'H': 24, 'I': 12, 'J': 8}
for c, w in col_w.items():
    ws1.column_dimensions[c].width = w

# ── Title Bar ──
r = 1
apply_fill(ws1, r, 1, 10, fill_navy)
ws1.merge_cells("B1:G1")
ws1["B1"].value = "Color of Money Decision Tree"
ws1["B1"].font = font_title
ws1["B1"].alignment = Alignment(horizontal="left", vertical="center")
ws1.merge_cells("H1:J1")
ws1["H1"].value = "ACQLERATE"
ws1["H1"].font = font_brand
ws1["H1"].alignment = Alignment(horizontal="right", vertical="center")
ws1.row_dimensions[1].height = 30

r = 2
apply_fill(ws1, r, 1, 10, fill_navy)
ws1.merge_cells("B2:G2")
ws1["B2"].value = "Appropriation Selection Guide — Defense Acquisitions Academy"
ws1["B2"].font = font_subtitle
ws1.merge_cells("H2:J2")
ws1["H2"].value = "Defense Acquisitions Academy"
ws1["H2"].font = Font(name="Calibri", size=8, color="B0C4DE")
ws1["H2"].alignment = Alignment(horizontal="right", vertical="center")
ws1.row_dimensions[2].height = 16

# ── Big Question ──
r = 4
ws1.merge_cells(f"B{r}:J{r}")
ws1.cell(row=r, column=2).value = '▼  "WHAT ARE YOU BUYING?"  — Start here to determine the correct appropriation'
ws1.cell(row=r, column=2).font = font_big_q
ws1.cell(row=r, column=2).alignment = ac
ws1.cell(row=r, column=2).fill = fill_light_teal
apply_fill(ws1, r, 2, 10, fill_light_teal)
ws1.row_dimensions[r].height = 28

# ── Column Headers for branches ──
r = 6
branch_headers = ["", "IF YOU'RE BUYING...", "APPROPRIATION", "OBLIGATION\nPERIOD", "COMMON USES", "COMMON MISTAKES", "ADA\nRISK"]
branch_cols = [(2,2), (3,3), (4,4), (5,5), (6,7), (8,8), (9,9)]
# Removed J from main but we'll use it for sub-info

branch_cols_full = [(2,2), (3,3), (4,4), (5,5), (6,7), (8,9), (10,10)]
for text, (c1, c2) in zip(branch_headers, branch_cols_full):
    ws1.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws1.cell(row=r, column=c1)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_teal
    cell.alignment = ac
    for c in range(c1, c2+1):
        ws1.cell(row=r, column=c).fill = fill_teal
        ws1.cell(row=r, column=c).border = thin
ws1.row_dimensions[r].height = 24

# ── Branch Data ──
branches = [
    {
        "icon": "👤",
        "buying": "PERSONNEL\n(Military pay, allowances,\nbonuses, PCS moves)",
        "approp": "MILPERS\n(Military Personnel)",
        "period": "1 Year",
        "uses": "Base pay, BAH, BAS, special pay, separation pay, PCS travel",
        "mistakes": "Using MILPERS for civilian pay (should be O&M); confusing with MILCON",
        "ada_risk": "M",
        "color": COLOR_MILPERS,
        "light": LIGHT_MILPERS,
    },
    {
        "icon": "🏗️",
        "buying": "BUILDINGS /\nCONSTRUCTION\n(Real property, facilities)",
        "approp": "MILCON\n(Military Construction)",
        "period": "5 Years",
        "uses": "New buildings, major renovations (>$1M), family housing, base infrastructure",
        "mistakes": "Using O&M for projects >$1M (threshold violation); not distinguishing from SRM",
        "ada_risk": "H",
        "color": COLOR_MILCON,
        "light": LIGHT_MILCON,
    },
    {
        "icon": "🔬",
        "buying": "RESEARCH &\nDEVELOPMENT\n(New tech, testing, R&D)",
        "approp": "RDT&E\n(Research, Dev, Test & Eval)",
        "period": "2 Years",
        "uses": "Basic/applied research, prototyping, system development, DT&E, OT&E, software dev",
        "mistakes": "Using RDT&E for production items; continuing R&D after LRIP decision (should be Proc)",
        "ada_risk": "M",
        "color": COLOR_RDTE,
        "light": LIGHT_RDTE,
    },
    {
        "icon": "⚙️",
        "buying": "EQUIPMENT / SYSTEMS\n(Investment items, major\nend items > $250K)",
        "approp": "PROCUREMENT\n(Various: Aircraft, Missile,\nShipbuilding, etc.)",
        "period": "3 Years",
        "uses": "Weapon systems, vehicles, aircraft, ships, communications equipment, spares (initial)",
        "mistakes": "Using O&M for items >$250K; splitting requirements to stay under threshold (ADA!)",
        "ada_risk": "H",
        "color": COLOR_PROC,
        "light": LIGHT_PROC,
    },
    {
        "icon": "🔧",
        "buying": "SERVICES / OPERATIONS\n(Day-to-day operations,\nmaintenance, supplies)",
        "approp": "O&M\n(Operations & Maintenance)",
        "period": "1 Year",
        "uses": "Civilian pay, travel, supplies <$250K, equipment maintenance, IT services, training",
        "mistakes": "Using O&M for investment items >$250K; paying for construction >$1M SRM threshold",
        "ada_risk": "H",
        "color": COLOR_OM,
        "light": LIGHT_OM,
    },
]

r = 7
for branch in branches:
    row_height = 48
    c_data = [
        (branch["icon"], (2, 2)),
        (branch["buying"], (3, 3)),
        (branch["approp"], (4, 4)),
        (branch["period"], (5, 5)),
        (branch["uses"], (6, 7)),
        (branch["mistakes"], (8, 9)),
        (branch["ada_risk"], (10, 10)),
    ]
    
    light_fill = PatternFill("solid", fgColor=branch["light"])
    approp_fill = PatternFill("solid", fgColor=branch["color"])
    
    for val, (c1, c2) in c_data:
        ws1.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws1.cell(row=r, column=c1)
        cell.value = val
        cell.alignment = ac
        cell.border = thin
        for c in range(c1, c2+1):
            ws1.cell(row=r, column=c).border = thin
        
        if c1 == 4:  # Appropriation column
            cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            cell.fill = approp_fill
            for c in range(c1, c2+1):
                ws1.cell(row=r, column=c).fill = approp_fill
        elif c1 == 10:  # ADA Risk
            risk_fonts = {
                "H": Font(name="Calibri", size=10, bold=True, color="C0392B"),
                "M": Font(name="Calibri", size=10, bold=True, color="D4850A"),
                "L": Font(name="Calibri", size=10, bold=True, color="27AE60"),
            }
            cell.font = risk_fonts.get(val, font_body)
            cell.fill = light_fill
            for c in range(c1, c2+1):
                ws1.cell(row=r, column=c).fill = light_fill
        elif c1 == 2:  # Icon
            cell.font = Font(name="Calibri", size=14)
            cell.fill = light_fill
            for c in range(c1, c2+1):
                ws1.cell(row=r, column=c).fill = light_fill
        else:
            cell.font = font_body if c1 > 5 else font_body_bold
            cell.fill = light_fill
            for c in range(c1, c2+1):
                ws1.cell(row=r, column=c).fill = light_fill
    
    ws1.row_dimensions[r].height = row_height
    r += 1

# ── RDT&E Budget Activity sub-table ──
r += 1
ws1.merge_cells(f"B{r}:J{r}")
ws1.cell(row=r, column=2).value = "RDT&E BUDGET ACTIVITIES (6.1 – 6.7)"
ws1.cell(row=r, column=2).font = font_section
ws1.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_RDTE)
ws1.cell(row=r, column=2).alignment = ac
apply_fill(ws1, r, 2, 10, PatternFill("solid", fgColor=COLOR_RDTE))
ws1.row_dimensions[r].height = 20
r += 1

rdte_headers = ["BA Code", "Category", "Description", "Examples"]
rdte_cols = [(2, 3), (4, 5), (6, 8), (9, 10)]
for text, (c1, c2) in zip(rdte_headers, rdte_cols):
    ws1.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws1.cell(row=r, column=c1)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_teal
    cell.alignment = ac
    for c in range(c1, c2+1):
        ws1.cell(row=r, column=c).fill = fill_teal
        ws1.cell(row=r, column=c).border = thin
ws1.row_dimensions[r].height = 18
r += 1

rdte_data = [
    ("6.1", "Basic Research", "Scientific study — discover new knowledge; not directed at specific military need", "University grants, lab research"),
    ("6.2", "Applied Research", "Translate basic research into solutions for military problems", "DARPA projects, concept demos"),
    ("6.3", "Advanced Tech Dev", "Develop & demonstrate technology in a relevant environment (prototyping)", "X-planes, ACAT III prototypes"),
    ("6.4", "Adv. Component Dev & Proto", "Evaluate integrated system concepts in operationally realistic conditions", "EMD phase pre-LRIP"),
    ("6.5", "System Dev & Demo", "Engineering & manufacturing dev to produce a system for operational testing", "Full system EMD, DT&E/OT&E"),
    ("6.6", "RDT&E Management Support", "Test ranges, evaluation facilities, R&D support infrastructure", "Test center ops, FFRDC support"),
    ("6.7", "Operational System Dev", "R&D for fielded systems — upgrades, modifications, software updates", "Block upgrades, tech insertion"),
]

for i, (ba, cat, desc, ex) in enumerate(rdte_data):
    alt = PatternFill("solid", fgColor=LIGHT_RDTE) if i % 2 == 0 else fill_white
    vals = [ba, cat, desc, ex]
    for val, (c1, c2) in zip(vals, rdte_cols):
        ws1.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws1.cell(row=r, column=c1)
        cell.value = val
        cell.font = font_body_bold if c1 <= 3 else font_body
        cell.alignment = al_indent if c1 > 3 else ac
        cell.fill = alt
        for c in range(c1, c2+1):
            ws1.cell(row=r, column=c).fill = alt
            ws1.cell(row=r, column=c).border = thin
    ws1.row_dimensions[r].height = 28
    r += 1

# ── Procurement threshold note ──
r += 1
ws1.merge_cells(f"B{r}:J{r}")
ws1.cell(row=r, column=2).value = '★ KEY THRESHOLD: Equipment/systems >$250K unit cost = Procurement. ≤$250K = consider O&M. Splitting requirements to stay under threshold is an ADA violation!'
ws1.cell(row=r, column=2).font = font_red_bold
ws1.cell(row=r, column=2).alignment = al_indent
ws1.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFF0F0")
apply_fill(ws1, r, 2, 10, PatternFill("solid", fgColor="FFF0F0"))
ws1.row_dimensions[r].height = 20

r += 1
ws1.merge_cells(f"B{r}:J{r}")
ws1.cell(row=r, column=2).value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com  |  ADA = Antideficiency Act (31 U.S.C. §1341)"
ws1.cell(row=r, column=2).font = font_footer
ws1.cell(row=r, column=2).alignment = ac

ws1.print_area = f"A1:J{r}"

# ═══════════════════════════════════════════════════
# SHEET 2: QUICK REFERENCE TABLE
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Quick Reference Table")
ws2.sheet_view.showGridLines = False

ws2.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws2.page_setup.orientation = "landscape"
ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.page_setup.fitToHeight = 1
ws2.page_setup.fitToWidth = 1
ws2.page_margins.left = 0.3
ws2.page_margins.right = 0.3
ws2.page_margins.top = 0.3
ws2.page_margins.bottom = 0.3

# Column widths
widths2 = {'A': 1.5, 'B': 13, 'C': 22, 'D': 11, 'E': 22, 'F': 22, 'G': 18, 'H': 12}
for c, w in widths2.items():
    ws2.column_dimensions[c].width = w

# Title
r = 1
apply_fill(ws2, r, 1, 8, fill_navy)
ws2.merge_cells("B1:F1")
ws2["B1"].value = "Color of Money — Quick Reference Table"
ws2["B1"].font = font_title
ws2["B1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.merge_cells("G1:H1")
ws2["G1"].value = "ACQLERATE"
ws2["G1"].font = font_brand
ws2["G1"].alignment = Alignment(horizontal="right", vertical="center")
ws2.row_dimensions[1].height = 30

r = 2
apply_fill(ws2, r, 1, 8, fill_navy)
ws2.merge_cells("B2:H2")
ws2["B2"].value = "Appropriation Types, Periods & Usage — Defense Acquisitions Academy"
ws2["B2"].font = font_subtitle
ws2.row_dimensions[2].height = 16

# Table headers
r = 4
ref_headers = ["Appropriation", "Full Name", "Obligation\nPeriod", "Best For", "Cannot Be Used For", "Common Example", "Key Threshold"]
ref_cols = [(2,2), (3,3), (4,4), (5,5), (6,6), (7,7), (8,8)]
for text, (c1, c2) in zip(ref_headers, ref_cols):
    cell = ws2.cell(row=r, column=c1)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_teal
    cell.alignment = ac
    cell.border = thin
ws2.row_dimensions[r].height = 28

# Data rows
ref_data = [
    ("O&M", "Operations & Maintenance", "1 Year", "Day-to-day ops, supplies <$250K, civilian pay, travel, maintenance, training, IT services",
     "Investment items >$250K, construction >$1M, military pay, R&D activities",
     "Office supplies, TDY travel, equipment repair", "$250K expense/investment threshold",
     COLOR_OM, LIGHT_OM),
    ("Procurement", "Various (Aircraft, Missile, Weapons, Shipbuilding, Other)", "3 Years",
     "Weapon systems, vehicles, aircraft, ships, communications equipment, initial spares",
     "R&D activities, O&M-type services, military pay, construction",
     "F-35 production lots, JLTV purchases", "$250K unit cost threshold",
     COLOR_PROC, LIGHT_PROC),
    ("RDT&E (6.1)", "Basic Research", "2 Years", "Scientific study — new knowledge discovery",
     "Production, operations, construction", "University research grants", "N/A — broadest flexibility",
     COLOR_RDTE, LIGHT_RDTE),
    ("RDT&E (6.2)", "Applied Research", "2 Years", "Applied R&D directed at military problems",
     "Production, operations", "DARPA technology programs", "N/A",
     COLOR_RDTE, LIGHT_RDTE),
    ("RDT&E (6.3)", "Advanced Technology Development", "2 Years", "Prototyping in relevant environment",
     "Production quantity buys", "Prototype demonstrations", "TRL 4-6 typical",
     COLOR_RDTE, LIGHT_RDTE),
    ("RDT&E (6.4)", "Adv. Component Dev & Prototypes", "2 Years", "Integrated system evaluation, pre-LRIP",
     "Full-rate production", "EMD contracts", "Milestone B decision",
     COLOR_RDTE, LIGHT_RDTE),
    ("RDT&E (6.5)", "System Development & Demonstration", "2 Years", "Full EMD, DT/OT, LRIP authorization",
     "Full-rate production (use Proc)", "F-35 SDD phase", "Milestone C triggers Proc",
     COLOR_RDTE, LIGHT_RDTE),
    ("RDT&E (6.6)", "RDT&E Management Support", "2 Years", "Test ranges, R&D infrastructure support",
     "Production, direct ops", "Test center operations", "Must be R&D-related",
     COLOR_RDTE, LIGHT_RDTE),
    ("RDT&E (6.7)", "Operational System Development", "2 Years", "Upgrades to fielded systems, tech insertion",
     "New-start programs (use 6.1-6.5)", "Software block upgrades", "Must be modification, not new start",
     COLOR_RDTE, LIGHT_RDTE),
    ("MILPERS", "Military Personnel", "1 Year", "Military pay, BAH, BAS, PCS, bonuses, separation pay",
     "Civilian salaries, equipment, construction, R&D",
     "Active duty monthly pay", "Only for uniformed personnel",
     COLOR_MILPERS, LIGHT_MILPERS),
    ("MILCON", "Military Construction", "5 Years", "New buildings, major renovations >$1M, family housing",
     "Sustainment/restoration <$1M (use O&M), equipment, pay",
     "New barracks, runway extension", "$1M SRM threshold / $6M unspecified minor threshold",
     COLOR_MILCON, LIGHT_MILCON),
    ("WCF", "Working Capital Fund", "No-Year\n(revolving)", "Depot maintenance, supply chain, transportation, information services",
     "Direct appropriation-type expenses outside WCF charter",
     "DLA supply purchases, depot-level repair", "Must break even over time",
     COLOR_WCF, LIGHT_WCF),
    ("Revolving Fund", "Various Revolving Funds", "No-Year\n(revolving)", "Self-sustaining commercial-type activities (commissary, AAFES support)",
     "Appropriated fund purposes", "DECA operations", "Revenue must cover costs",
     COLOR_REVFUND, LIGHT_REVFUND),
]

r = 5
for (approp, full, period, best, cannot, example, threshold, color, light) in ref_data:
    vals = [approp, full, period, best, cannot, example, threshold]
    approp_fill = PatternFill("solid", fgColor=color)
    light_fill = PatternFill("solid", fgColor=light)
    
    for i, (val, (c1, c2)) in enumerate(zip(vals, ref_cols)):
        cell = ws2.cell(row=r, column=c1)
        cell.value = val
        cell.border = thin
        
        if i == 0:
            cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            cell.fill = approp_fill
            cell.alignment = ac
        else:
            cell.font = font_body
            cell.fill = light_fill
            cell.alignment = al_indent if i in [3, 4] else ac
    
    ws2.row_dimensions[r].height = 40
    r += 1

# Footer
r += 1
ws2.merge_cells(f"B{r}:H{r}")
ws2.cell(row=r, column=2).value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com  |  Ref: DoD 7000.14-R, FMR Vol 2A/3; FAR Part 31; 31 U.S.C. §§1301-1558"
ws2.cell(row=r, column=2).font = font_footer
ws2.cell(row=r, column=2).alignment = ac

ws2.print_area = f"A1:H{r}"

# ── Save ──
output = "/home/user/workspace/acq-pro/products/pack3-finance-cheat-sheets/color-of-money-decision-tree.xlsx"
wb.save(output)
print(f"Saved: {output}")
