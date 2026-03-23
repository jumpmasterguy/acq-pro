"""
FILE 3: evm-formulas-quick-reference.xlsx
EVM Quick Reference Card — single beautiful print-ready sheet
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EVM Quick Reference"

# ── Brand colors ──
TEAL = "01696F"
NAVY = "0D1B2A"
WHITE = "FFFFFF"
LIGHT_TEAL = "E0F2F3"
VERY_LIGHT_TEAL = "F0FAFA"
DARK_GRAY = "333333"
MID_GRAY = "666666"
GREEN = "1B7A3D"
RED = "C0392B"
AMBER = "D4850A"
LIGHT_GREEN = "E8F5E9"
LIGHT_RED = "FDEDEC"
LIGHT_AMBER = "FFF8E1"
BORDER_GRAY = "C0C0C0"
LIGHT_GRAY = "F5F5F5"

# ── Fills ──
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_teal = PatternFill("solid", fgColor=TEAL)
fill_light_teal = PatternFill("solid", fgColor=LIGHT_TEAL)
fill_very_light_teal = PatternFill("solid", fgColor=VERY_LIGHT_TEAL)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_light_gray = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_green = PatternFill("solid", fgColor=LIGHT_GREEN)
fill_red = PatternFill("solid", fgColor=LIGHT_RED)
fill_amber = PatternFill("solid", fgColor=LIGHT_AMBER)

# ── Fonts ──
font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=9, color="B0C4DE")
font_brand = Font(name="Calibri", size=12, bold=True, color="4DA8AE")
font_section = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_subsection = Font(name="Calibri", size=9, bold=True, color=TEAL)
font_header = Font(name="Calibri", size=8, bold=True, color=WHITE)
font_body = Font(name="Calibri", size=8, color=DARK_GRAY)
font_body_bold = Font(name="Calibri", size=8, bold=True, color=DARK_GRAY)
font_formula = Font(name="Consolas", size=9, bold=True, color=TEAL)
font_formula_small = Font(name="Consolas", size=8, color=TEAL)
font_green = Font(name="Calibri", size=8, color=GREEN)
font_red = Font(name="Calibri", size=8, color=RED)
font_amber = Font(name="Calibri", size=8, color=AMBER)
font_footer = Font(name="Calibri", size=7, italic=True, color=MID_GRAY)
font_term = Font(name="Calibri", size=7, bold=True, color=TEAL)
font_term_def = Font(name="Calibri", size=7, color=MID_GRAY)

# ── Borders ──
thin = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

# ── Alignments ──
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
al_indent = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

# ── Page Setup ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3
ws.sheet_view.showGridLines = False

# ── Column widths (A-I, 9 cols) ──
widths = {'A': 1.5, 'B': 14, 'C': 22, 'D': 16, 'E': 16, 'F': 16, 'G': 14, 'H': 14, 'I': 14}
for c, w in widths.items():
    ws.column_dimensions[c].width = w

def apply_row(ws, row, col_start, col_end, fill, border=None):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
        if border:
            ws.cell(row=row, column=c).border = border

def section_header(ws, row, text, col_start=2, col_end=9):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = font_section
    cell.fill = fill_navy
    cell.alignment = ac
    apply_row(ws, row, col_start, col_end, fill_navy)
    ws.row_dimensions[row].height = 20
    return row + 1

def table_header_row(ws, row, headers, col_ranges):
    for text, (c1, c2) in zip(headers, col_ranges):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = text
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = ac
        for c in range(c1, c2+1):
            ws.cell(row=row, column=c).fill = fill_teal
            ws.cell(row=row, column=c).border = thin
    ws.row_dimensions[row].height = 16
    return row + 1

def data_row(ws, row, values, col_ranges, fonts=None, fills=None, height=18):
    default_font = font_body
    default_fill = fill_white
    for i, (text, (c1, c2)) in enumerate(zip(values, col_ranges)):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = text
        cell.font = fonts[i] if fonts else default_font
        cell.fill = fills[i] if fills else default_fill
        cell.alignment = al_indent if c1 == 2 else ac
        for c in range(c1, c2+1):
            ws.cell(row=row, column=c).fill = fills[i] if fills else default_fill
            ws.cell(row=row, column=c).border = thin
    ws.row_dimensions[row].height = height
    return row + 1

# ═══════════════════════════════════════════════════
# TITLE BAR
# ═══════════════════════════════════════════════════
r = 1
apply_row(ws, r, 1, 9, fill_navy)
ws.merge_cells("B1:F1")
ws["B1"].value = "EVM Formulas Quick Reference"
ws["B1"].font = font_title
ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("G1:I1")
ws["G1"].value = "ACQLERATE"
ws["G1"].font = font_brand
ws["G1"].alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[1].height = 28

r = 2
apply_row(ws, r, 1, 9, fill_navy)
ws.merge_cells("B2:F2")
ws["B2"].value = "Earned Value Management — Defense Acquisitions Academy"
ws["B2"].font = font_subtitle
ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("G2:I2")
ws["G2"].value = "Pin to your wall!"
ws["G2"].font = Font(name="Calibri", size=8, italic=True, color="B0C4DE")
ws["G2"].alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[2].height = 16

r = 3
ws.row_dimensions[3].height = 5

# ═══════════════════════════════════════════════════
# SECTION 1: THE THREE VALUES
# ═══════════════════════════════════════════════════
r = 4
r = section_header(ws, r, "1. THE THREE FUNDAMENTAL VALUES")

cols3 = [(2,3), (4,6), (7,9)]
r = table_header_row(ws, r, ["Metric", "Plain English", "Also Known As"], cols3)

three_values = [
    ("PV (Planned Value)", "The budgeted cost for work scheduled to be done by now", "BCWS — Budgeted Cost of Work Scheduled"),
    ("EV (Earned Value)", "The budgeted cost for work actually completed so far", "BCWP — Budgeted Cost of Work Performed"),
    ("AC (Actual Cost)", "The actual cost incurred for work completed so far", "ACWP — Actual Cost of Work Performed"),
]
for val_name, desc, aka in three_values:
    alt = fill_light_teal if (r % 2 == 0) else fill_white
    r = data_row(ws, r, [val_name, desc, aka], cols3,
                 fonts=[font_body_bold, font_body, font_body],
                 fills=[alt, alt, alt])

# ═══════════════════════════════════════════════════
# SECTION 2: VARIANCE METRICS
# ═══════════════════════════════════════════════════
r += 0
r = section_header(ws, r, "2. VARIANCE METRICS")

cols4 = [(2,3), (4,5), (6,7), (8,9)]
r = table_header_row(ws, r, ["Metric", "Formula", "Positive = Good", "Negative = Bad"], cols4)

variances = [
    ("CV (Cost Variance)", "EV − AC", "Under budget", "Over budget"),
    ("SV (Schedule Variance)", "EV − PV", "Ahead of schedule", "Behind schedule"),
    ("VAC (Variance at Completion)", "BAC − EAC", "Expected underrun", "Expected overrun"),
]
for name, formula, pos, neg in variances:
    alt = fill_light_teal if (r % 2 == 0) else fill_white
    r = data_row(ws, r, [name, formula, pos, neg], cols4,
                 fonts=[font_body_bold, font_formula_small, font_green, font_red],
                 fills=[alt, alt, fill_green, fill_red], height=18)

# ═══════════════════════════════════════════════════
# SECTION 3: INDEX METRICS
# ═══════════════════════════════════════════════════
r = section_header(ws, r, "3. PERFORMANCE INDEX METRICS")

cols_idx = [(2,3), (4,5), (6,7), (8,9)]
r = table_header_row(ws, r, ["Index", "Formula", "> 1.0", "< 1.0"], cols_idx)

indices = [
    ("CPI (Cost Performance Index)", "EV / AC", "Under budget — efficient", "Over budget — burning cash"),
    ("SPI (Schedule Perf. Index)", "EV / PV", "Ahead of schedule", "Behind schedule"),
    ("TCPI (To-Complete Perf. Index)", "(BAC−EV) / (BAC−AC)", "Lower = easier path forward", "Higher = harder to recover"),
]
for name, formula, good, bad in indices:
    alt = fill_light_teal if (r % 2 == 0) else fill_white
    r = data_row(ws, r, [name, formula, good, bad], cols_idx,
                 fonts=[font_body_bold, font_formula_small, font_green, font_red],
                 fills=[alt, alt, fill_green, fill_red], height=18)

# CPI rule note
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws.cell(row=r, column=2).value = "★ Rule of Thumb: After 20% completion, CPI rarely improves by more than 10%. Plan accordingly — history shows overruns persist."
ws.cell(row=r, column=2).font = Font(name="Calibri", size=7.5, italic=True, bold=True, color=RED)
ws.cell(row=r, column=2).alignment = al_indent
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFF0F0")
apply_row(ws, r, 2, 9, PatternFill("solid", fgColor="FFF0F0"))
ws.row_dimensions[r].height = 16
r += 1

# ═══════════════════════════════════════════════════
# SECTION 4: EAC FORMULAS
# ═══════════════════════════════════════════════════
r = section_header(ws, r, "4. ESTIMATE AT COMPLETION (EAC) — Four Methods")

cols_eac = [(2,2), (3,4), (5,7), (8,9)]
r = table_header_row(ws, r, ["Method", "Formula", "When to Use", "Accuracy"], cols_eac)

eac_methods = [
    ("#1 — CPI-Based", "BAC / CPI", "Most common. Assumes future cost efficiency = past. Most accurate after 20% complete.", "★★★★"),
    ("#2 — Optimistic", "AC + (BAC − EV)", "Assumes overrun is non-recurring (one-time issue). Use with caution.", "★★"),
    ("#3 — PM Estimate", "AC + ETC (new est.)", "PM provides fresh bottom-up estimate of remaining work. Use for major re-plans.", "★★★"),
    ("#4 — Pessimistic", "AC + (BAC−EV)/(CPI×SPI)", "Accounts for BOTH cost and schedule inefficiency. Worst-case realism.", "★★★★"),
]
for method, formula, when, accuracy in eac_methods:
    alt = fill_light_teal if (r % 2 == 0) else fill_white
    r = data_row(ws, r, [method, formula, when, accuracy], cols_eac,
                 fonts=[font_body_bold, font_formula_small, font_body, font_body_bold],
                 fills=[alt, alt, alt, alt], height=30)

# ═══════════════════════════════════════════════════
# SECTION 5: QUICK INTERPRETATION GUIDE (Traffic Light)
# ═══════════════════════════════════════════════════
r = section_header(ws, r, "5. QUICK INTERPRETATION GUIDE — CPI / SPI Ranges")

cols_tl = [(2,3), (4,5), (6,7), (8,9)]
r = table_header_row(ws, r, ["Range", "Status", "What It Means", "Action Required"], cols_tl)

traffic = [
    ("≥ 1.00", "GREEN", "On or under budget/schedule", "Continue monitoring", fill_green, font_green),
    ("0.90 – 0.99", "AMBER", "Slightly over budget/behind schedule", "Investigate root cause; corrective action plan", fill_amber, font_amber),
    ("0.80 – 0.89", "RED", "Significant cost/schedule overrun", "Formal CAP required; OTB/OTS consideration", fill_red, font_red),
    ("< 0.80", "CRITICAL", "Severe — likely unrecoverable without re-baseline", "EAC revision; Nunn-McCurdy breach possible (>15%)", fill_red, Font(name="Calibri", size=8, bold=True, color=RED)),
]
for rng, status, meaning, action, fill, status_font in traffic:
    r = data_row(ws, r, [rng, status, meaning, action], cols_tl,
                 fonts=[font_body_bold, status_font, font_body, font_body],
                 fills=[fill_white, fill, fill_white, fill_white], height=20)

# ═══════════════════════════════════════════════════
# SECTION 6: CPR FORMAT REFERENCE
# ═══════════════════════════════════════════════════
r = section_header(ws, r, "6. CONTRACT PERFORMANCE REPORT (CPR) FORMATS")

cols_cpr = [(2,3), (4,9)]
r = table_header_row(ws, r, ["Format", "Description"], cols_cpr)

cpr_formats = [
    ("Format 1 — WBS", "Work Breakdown Structure — cost/schedule data by WBS element (product-oriented)"),
    ("Format 2 — Org", "Organizational — cost/schedule data by performing organization (functional view)"),
    ("Format 3 — Baseline", "Baseline — time-phased PMB showing budget distribution over time"),
    ("Format 4 — Staffing", "Staffing — labor hours (actual & forecast) by month, by functional category"),
    ("Format 5 — Narrative", "Explanations & Variance Analysis — written analysis of significant variances"),
]
for fmt, desc in cpr_formats:
    alt = fill_light_teal if (r % 2 == 0) else fill_white
    r = data_row(ws, r, [fmt, desc], cols_cpr,
                 fonts=[font_body_bold, font_body],
                 fills=[alt, alt], height=16)

# ═══════════════════════════════════════════════════
# SECTION 7: KEY TERMS
# ═══════════════════════════════════════════════════
r = section_header(ws, r, "7. KEY EVM TERMS")

terms = [
    ("PMB", "Performance Measurement Baseline — the time-phased budget plan against which performance is measured"),
    ("BAC", "Budget At Completion — total budget for the authorized work (PMB total)"),
    ("CBB", "Contract Budget Base — negotiated contract cost + authorized unpriced work"),
    ("MR", "Management Reserve — budget held by PM for unknown unknowns; NOT in PMB"),
    ("UB", "Undistributed Budget — budget allocated to the contract but not yet assigned to WBS elements"),
    ("WBS", "Work Breakdown Structure — product-oriented hierarchy decomposing project scope"),
    ("CAP", "Control Account Plan — the management control point where scope, schedule, budget intersect"),
    ("OTB/OTS", "Over Target Baseline/Schedule — formal re-baselining when PMB is no longer achievable"),
]

# 2-col layout for terms
cols_t1 = [(2, 2), (3, 5)]
cols_t2 = [(6, 6), (7, 9)]

for i in range(0, len(terms), 2):
    # Left term
    acronym, defn = terms[i]
    ws.cell(row=r, column=2).value = acronym
    ws.cell(row=r, column=2).font = font_term
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="top")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(row=r, column=3).value = defn
    ws.cell(row=r, column=3).font = font_term_def
    ws.cell(row=r, column=3).alignment = al_top
    
    # Right term (if exists)
    if i + 1 < len(terms):
        acronym2, defn2 = terms[i+1]
        ws.cell(row=r, column=6).value = acronym2
        ws.cell(row=r, column=6).font = font_term
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="top")
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        ws.cell(row=r, column=7).value = defn2
        ws.cell(row=r, column=7).font = font_term_def
        ws.cell(row=r, column=7).alignment = al_top
    
    ws.row_dimensions[r].height = 22
    r += 1

# ═══════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws.cell(row=r, column=2).value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com  |  Reference: ANSI/EIA-748, DoDI 5000.02, OMB Circular A-11"
ws.cell(row=r, column=2).font = font_footer
ws.cell(row=r, column=2).alignment = ac

# Print area
ws.print_area = f"A1:I{r}"

output = "/home/user/workspace/acq-pro/products/pack3-finance-cheat-sheets/evm-formulas-quick-reference.xlsx"
wb.save(output)
print(f"Saved: {output}")
