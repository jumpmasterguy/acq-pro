"""
FILE 4: wrap-rate-breakdown.xlsx
Wrap rate calculator, comparison, and instructions — 3 sheets
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

wb = openpyxl.Workbook()

# ── Brand colors ──
TEAL = "01696F"
NAVY = "0D1B2A"
WHITE = "FFFFFF"
LIGHT_TEAL = "E0F2F3"
VERY_LIGHT_TEAL = "F0FAFA"
DARK_GRAY = "333333"
MID_GRAY = "666666"
BORDER_GRAY = "C0C0C0"
LIGHT_GRAY = "F5F5F5"
INPUT_BLUE = "0000FF"  # Financial model convention
LIGHT_YELLOW = "FFFFF0"

# ── Fills ──
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_teal = PatternFill("solid", fgColor=TEAL)
fill_light_teal = PatternFill("solid", fgColor=LIGHT_TEAL)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_light_gray = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_very_light_teal = PatternFill("solid", fgColor=VERY_LIGHT_TEAL)
fill_input = PatternFill("solid", fgColor="FFFFF0")  # Light yellow for inputs
fill_output = PatternFill("solid", fgColor="E8F5E9")  # Light green for outputs
fill_result = PatternFill("solid", fgColor=LIGHT_TEAL)

# ── Fonts ──
font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=9, color="B0C4DE")
font_brand = Font(name="Calibri", size=12, bold=True, color="4DA8AE")
font_section = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_header = Font(name="Calibri", size=9, bold=True, color=WHITE)
font_body = Font(name="Calibri", size=10, color=DARK_GRAY)
font_body_bold = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
font_body_sm = Font(name="Calibri", size=9, color=DARK_GRAY)
font_input = Font(name="Calibri", size=11, bold=True, color=INPUT_BLUE)  # Blue = input
font_formula = Font(name="Calibri", size=11, color=DARK_GRAY)  # Black = formula
font_formula_bold = Font(name="Calibri", size=11, bold=True, color=DARK_GRAY)
font_result = Font(name="Calibri", size=13, bold=True, color=TEAL)
font_label = Font(name="Calibri", size=10, color=MID_GRAY)
font_label_bold = Font(name="Calibri", size=10, bold=True, color=TEAL)
font_note = Font(name="Calibri", size=8, italic=True, color=MID_GRAY)
font_footer = Font(name="Calibri", size=7, italic=True, color=MID_GRAY)
font_inst_body = Font(name="Calibri", size=10, color=DARK_GRAY)
font_inst_bold = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
font_inst_head = Font(name="Calibri", size=11, bold=True, color=TEAL)

# ── Borders ──
thin = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
bottom_thin = Border(bottom=Side(style="thin", color=BORDER_GRAY))
bottom_thick = Border(bottom=Side(style="medium", color=TEAL))
input_border = Border(
    left=Side(style="thin", color=INPUT_BLUE),
    right=Side(style="thin", color=INPUT_BLUE),
    top=Side(style="thin", color=INPUT_BLUE),
    bottom=Side(style="thin", color=INPUT_BLUE),
)

# ── Alignments ──
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_indent = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ar = Alignment(horizontal="right", vertical="center")
al_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

def apply_fill(ws, row, c1, c2, fill):
    for c in range(c1, c2+1):
        ws.cell(row=row, column=c).fill = fill

# ═══════════════════════════════════════════════════
# SHEET 1: WRAP RATE CALCULATOR
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Wrap Rate Calculator"
ws1.sheet_view.showGridLines = False

ws1.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws1.page_setup.orientation = "portrait"
ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
ws1.page_setup.fitToHeight = 1
ws1.page_setup.fitToWidth = 1
ws1.page_margins.left = 0.5
ws1.page_margins.right = 0.5
ws1.page_margins.top = 0.4
ws1.page_margins.bottom = 0.4

# Column widths: A=margin, B=labels, C=rates/values, D=formulas/notes, E=extra
widths = {'A': 2, 'B': 32, 'C': 16, 'D': 28, 'E': 16, 'F': 3}
for c, w in widths.items():
    ws1.column_dimensions[c].width = w

# ── Title Bar ──
r = 1
apply_fill(ws1, r, 1, 6, fill_navy)
ws1.merge_cells("B1:D1")
ws1["B1"].value = "Wrap Rate Calculator"
ws1["B1"].font = font_title
ws1["B1"].alignment = Alignment(horizontal="left", vertical="center")
ws1.merge_cells("E1:F1")
ws1["E1"].value = "ACQLERATE"
ws1["E1"].font = font_brand
ws1["E1"].alignment = Alignment(horizontal="right", vertical="center")
ws1.row_dimensions[1].height = 30

r = 2
apply_fill(ws1, r, 1, 6, fill_navy)
ws1.merge_cells("B2:D2")
ws1["B2"].value = "Defense Contractor Rate Buildup — Defense Acquisitions Academy"
ws1["B2"].font = font_subtitle
ws1.merge_cells("E2:F2")
ws1["E2"].value = "Blue text = inputs you can change"
ws1["E2"].font = Font(name="Calibri", size=8, italic=True, color="B0C4DE")
ws1["E2"].alignment = Alignment(horizontal="right", vertical="center")
ws1.row_dimensions[2].height = 16

# ── INPUT SECTION ──
r = 4
ws1.merge_cells(f"B{r}:E{r}")
ws1.cell(row=r, column=2).value = "INPUT ASSUMPTIONS"
ws1.cell(row=r, column=2).font = font_section
ws1.cell(row=r, column=2).fill = fill_teal
ws1.cell(row=r, column=2).alignment = ac
apply_fill(ws1, r, 2, 5, fill_teal)
ws1.row_dimensions[r].height = 22

r = 5
# Headers for input section
for col, hdr in [(2, "Input"), (3, "Value"), (4, "Description")]:
    cell = ws1.cell(row=r, column=col)
    cell.value = hdr
    cell.font = Font(name="Calibri", size=9, bold=True, color=TEAL)
    cell.alignment = ac
    cell.border = bottom_thin
ws1.row_dimensions[r].height = 18

# Input rows with example values
inputs = [
    ("Base Hourly Rate", 65, "$#,##0.00", "Direct labor hourly rate (loaded into labor category)"),
    ("Fringe Benefit Rate %", 0.30, "0.0%", "FICA, health insurance, 401k, PTO, workers comp"),
    ("Overhead Rate %", 0.45, "0.0%", "Indirect costs: mgmt, facilities, IT, HR (applied to DL + Fringe)"),
    ("G&A Rate %", 0.12, "0.0%", "General & Administrative: exec, legal, finance (applied to Total Cost Input)"),
    ("Fee / Profit %", 0.08, "0.0%", "Contractor profit margin (negotiated; varies by contract type)"),
]

r = 6
input_rows = {}
for label, value, fmt, desc in inputs:
    ws1.cell(row=r, column=2).value = label
    ws1.cell(row=r, column=2).font = font_body
    ws1.cell(row=r, column=2).alignment = al_indent
    
    cell = ws1.cell(row=r, column=3)
    cell.value = value
    cell.font = font_input  # Blue text = input
    cell.number_format = fmt
    cell.alignment = ac
    cell.fill = fill_input
    cell.border = input_border
    
    ws1.cell(row=r, column=4).value = desc
    ws1.cell(row=r, column=4).font = font_note
    ws1.cell(row=r, column=4).alignment = al
    
    input_rows[label] = r
    ws1.row_dimensions[r].height = 22
    r += 1

# Cell references for formulas
base_cell = f"C{input_rows['Base Hourly Rate']}"
fringe_pct = f"C{input_rows['Fringe Benefit Rate %']}"
oh_pct = f"C{input_rows['Overhead Rate %']}"
ga_pct = f"C{input_rows['G&A Rate %']}"
fee_pct = f"C{input_rows['Fee / Profit %']}"

# ── OUTPUT / BUILDUP SECTION ──
r += 1
ws1.merge_cells(f"B{r}:E{r}")
ws1.cell(row=r, column=2).value = "RATE BUILDUP (All formulas — change inputs above to recalculate)"
ws1.cell(row=r, column=2).font = font_section
ws1.cell(row=r, column=2).fill = fill_navy
ws1.cell(row=r, column=2).alignment = ac
apply_fill(ws1, r, 2, 5, fill_navy)
ws1.row_dimensions[r].height = 22
r += 1

# Column headers
for col, hdr in [(2, "Component"), (3, "Amount"), (4, "Formula"), (5, "Running Total")]:
    cell = ws1.cell(row=r, column=col)
    cell.value = hdr
    cell.font = Font(name="Calibri", size=9, bold=True, color=TEAL)
    cell.alignment = ac
    cell.border = bottom_thin
ws1.row_dimensions[r].height = 18
r += 1

# Buildup rows
buildup = [
    ("Direct Labor (DL)", f"={base_cell}", "$#,##0.00",
     f'="$"&TEXT({base_cell},"#,##0.00")&"/hr (base rate)"',
     f"={base_cell}"),
    ("+ Fringe Benefits", f"={base_cell}*{fringe_pct}", "$#,##0.00",
     f'="DL × "&TEXT({fringe_pct},"0.0%")',
     f"={base_cell}+{base_cell}*{fringe_pct}"),
    ("\u2192 Fringe-Burdened Rate", f"={base_cell}*(1+{fringe_pct})", "$#,##0.00",
     '"DL + Fringe"',
     f"={base_cell}*(1+{fringe_pct})"),
    ("+ Overhead", f"={base_cell}*(1+{fringe_pct})*{oh_pct}", "$#,##0.00",
     f'="Fringe-Burdened × "&TEXT({oh_pct},"0.0%")',
     f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})"),
    ("\u2192 Total Cost Input (TCI)", f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})", "$#,##0.00",
     '"DL + Fringe + Overhead"',
     f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})"),
    ("+ G&A", f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*{ga_pct}", "$#,##0.00",
     f'="TCI × "&TEXT({ga_pct},"0.0%")',
     f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})"),
    ("\u2192 Total Cost", f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})", "$#,##0.00",
     '"TCI + G&A"',
     f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})"),
    ("+ Fee / Profit", f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})*{fee_pct}", "$#,##0.00",
     f'="Total Cost × "&TEXT({fee_pct},"0.0%")',
     f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})*(1+{fee_pct})"),
]

for label, formula, fmt, formula_desc, running in buildup:
    is_subtotal = label.startswith("\u2192")
    is_final = "Total Cost" in label and "Input" not in label and "+" not in label
    
    ws1.cell(row=r, column=2).value = label
    ws1.cell(row=r, column=2).font = font_body_bold if is_subtotal else font_body
    ws1.cell(row=r, column=2).alignment = al_indent
    
    cell_c = ws1.cell(row=r, column=3)
    cell_c.value = formula
    cell_c.font = font_formula_bold if is_subtotal else font_formula
    cell_c.number_format = fmt
    cell_c.alignment = ac
    cell_c.fill = fill_output if is_subtotal else fill_white
    
    ws1.cell(row=r, column=4).value = f"={formula_desc}" if not formula_desc.startswith("=") else formula_desc
    ws1.cell(row=r, column=4).font = font_note
    ws1.cell(row=r, column=4).alignment = al
    
    cell_e = ws1.cell(row=r, column=5)
    cell_e.value = running
    cell_e.font = font_formula_bold if is_subtotal else font_formula
    cell_e.number_format = fmt
    cell_e.alignment = ac
    cell_e.fill = fill_output if is_subtotal else fill_white
    
    if is_subtotal:
        for c in range(2, 6):
            ws1.cell(row=r, column=c).border = Border(
                top=Side(style="thin", color=BORDER_GRAY),
                bottom=Side(style="thin", color=BORDER_GRAY),
            )
    
    ws1.row_dimensions[r].height = 22
    r += 1

# ── FINAL RESULTS ──
r += 1
ws1.merge_cells(f"B{r}:E{r}")
ws1.cell(row=r, column=2).value = "FINAL RESULTS"
ws1.cell(row=r, column=2).font = font_section
ws1.cell(row=r, column=2).fill = fill_teal
ws1.cell(row=r, column=2).alignment = ac
apply_fill(ws1, r, 2, 5, fill_teal)
ws1.row_dimensions[r].height = 22
r += 1

# Fully Loaded Billing Rate
ws1.cell(row=r, column=2).value = "Fully Loaded Billing Rate"
ws1.cell(row=r, column=2).font = font_label_bold
ws1.cell(row=r, column=2).alignment = al_indent

cell = ws1.cell(row=r, column=3)
cell.value = f"={base_cell}*(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})*(1+{fee_pct})"
cell.font = font_result
cell.number_format = "$#,##0.00"
cell.alignment = ac
cell.fill = fill_result
cell.border = Border(
    left=Side(style="medium", color=TEAL),
    right=Side(style="medium", color=TEAL),
    top=Side(style="medium", color=TEAL),
    bottom=Side(style="medium", color=TEAL),
)

ws1.merge_cells(f"D{r}:E{r}")
ws1.cell(row=r, column=4).value = "Total Cost + Fee = what the government pays per hour"
ws1.cell(row=r, column=4).font = font_note
ws1.cell(row=r, column=4).alignment = al
ws1.row_dimensions[r].height = 28
r += 1

# Wrap Rate Multiplier
ws1.cell(row=r, column=2).value = "Wrap Rate Multiplier"
ws1.cell(row=r, column=2).font = font_label_bold
ws1.cell(row=r, column=2).alignment = al_indent

cell = ws1.cell(row=r, column=3)
cell.value = f"=(1+{fringe_pct})*(1+{oh_pct})*(1+{ga_pct})*(1+{fee_pct})"
cell.font = font_result
cell.number_format = "0.00x"
cell.alignment = ac
cell.fill = fill_result
cell.border = Border(
    left=Side(style="medium", color=TEAL),
    right=Side(style="medium", color=TEAL),
    top=Side(style="medium", color=TEAL),
    bottom=Side(style="medium", color=TEAL),
)

ws1.merge_cells(f"D{r}:E{r}")
ws1.cell(row=r, column=4).value = "Fully Loaded Rate ÷ Base Rate (how many times base the govt pays)"
ws1.cell(row=r, column=4).font = font_note
ws1.cell(row=r, column=4).alignment = al
ws1.row_dimensions[r].height = 28
r += 2

# Color legend
ws1.merge_cells(f"B{r}:E{r}")
ws1.cell(row=r, column=2).value = "★ Convention: Blue text = hardcoded inputs (change these)  |  Black text = formulas (auto-calculated)  |  Yellow background = input cells  |  Green = subtotals"
ws1.cell(row=r, column=2).font = font_note
ws1.cell(row=r, column=2).alignment = al
ws1.row_dimensions[r].height = 16
r += 1

ws1.merge_cells(f"B{r}:E{r}")
ws1.cell(row=r, column=2).value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com"
ws1.cell(row=r, column=2).font = font_footer
ws1.cell(row=r, column=2).alignment = ac
ws1.print_area = f"A1:F{r}"

# ═══════════════════════════════════════════════════
# SHEET 2: RATE COMPARISON
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Rate Comparison")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws2.page_setup.orientation = "landscape"
ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.page_setup.fitToHeight = 1
ws2.page_setup.fitToWidth = 1
ws2.page_margins.left = 0.4
ws2.page_margins.right = 0.4
ws2.page_margins.top = 0.4
ws2.page_margins.bottom = 0.4

widths2 = {'A': 2, 'B': 22, 'C': 13, 'D': 13, 'E': 13, 'F': 11, 'G': 14, 'H': 36}
for c, w in widths2.items():
    ws2.column_dimensions[c].width = w

# Title
r = 1
apply_fill(ws2, r, 1, 8, fill_navy)
ws2.merge_cells("B1:F1")
ws2["B1"].value = "Wrap Rate Comparison by Contractor Type"
ws2["B1"].font = font_title
ws2["B1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.merge_cells("G1:H1")
ws2["G1"].value = "ACQLERATE"
ws2["G1"].font = font_brand
ws2["G1"].alignment = Alignment(horizontal="right", vertical="center")
ws2.row_dimensions[1].height = 30

r = 2
apply_fill(ws2, r, 1, 8, fill_navy)
ws2.merge_cells("B2:H2")
ws2["B2"].value = "Typical rates across defense industry segments — Defense Acquisitions Academy"
ws2["B2"].font = font_subtitle
ws2.row_dimensions[2].height = 16

# Table headers
r = 4
comp_headers = ["Contractor Type", "Typical Fringe", "Typical OH", "Typical G&A", "Typical Fee", "Typical\nWrap Rate", "Notes"]
comp_cols = list(range(2, 9))
for i, (hdr, col) in enumerate(zip(comp_headers, comp_cols)):
    cell = ws2.cell(row=r, column=col)
    cell.value = hdr
    cell.font = font_header
    cell.fill = fill_teal
    cell.alignment = ac
    cell.border = thin
ws2.row_dimensions[r].height = 28

# Data
comp_data = [
    ("Large Prime\n(LM, RTX, NGC, BA, GD)", "32-38%", "50-70%", "10-15%", "8-10%", "2.8x – 3.5x",
     "Highest overhead due to IRAD, facilities, compliance infrastructure. DCAA-audited rates. Competitive on large ACAT I/II programs."),
    ("Mid-Size Defense\n(SAIC, Leidos, Booz Allen, CACI)", "28-35%", "40-55%", "10-14%", "7-10%", "2.3x – 2.9x",
     "Sweet spot for professional services. Lower OH than primes. Strong in A&AS, SETA, IT contracts."),
    ("Small Business\n(8(a), SDVOSB, HUBZone)", "22-30%", "30-45%", "8-15%", "8-15%", "1.9x – 2.5x",
     "Lower OH but potentially higher G&A (smaller base). Fee can be higher on sole-source. May lack DCAA-approved rates initially."),
    ("Specialty / Niche\n(Cleared professionals, cyber)", "30-38%", "45-65%", "10-16%", "8-12%", "2.5x – 3.2x",
     "High fringe due to cleared workforce retention. Premium for specialized skills. Cyber and SIGINT talent commands premium."),
    ("OCONUS / Deployed\n(Overseas operations)", "35-45%", "55-80%", "12-18%", "8-12%", "3.0x – 4.0x",
     "Hazard/hardship differentials, LOGCAP support costs, security, housing, R&R travel. OCONUS uplift can be 1.3-1.8x CONUS rates."),
    ("FFRDC / UARC\n(MITRE, IDA, Lincoln Labs)", "35-40%", "55-75%", "12-16%", "0% (non-profit)", "2.4x – 3.0x",
     "No fee (non-profit). High OH reflects IRAD equivalent, facilities. Exempt from some cost principles. Sole-source authority."),
]

r = 5
for i, (ctype, fringe, oh, ga, fee, wrap, notes) in enumerate(comp_data):
    alt = fill_light_teal if i % 2 == 0 else fill_white
    vals = [ctype, fringe, oh, ga, fee, wrap, notes]
    for j, (val, col) in enumerate(zip(vals, comp_cols)):
        cell = ws2.cell(row=r, column=col)
        cell.value = val
        cell.border = thin
        cell.fill = alt
        if j == 0:
            cell.font = font_body_bold
            cell.alignment = al_indent
        elif j == 6:
            cell.font = Font(name="Calibri", size=8, color=MID_GRAY)
            cell.alignment = al_top
        else:
            cell.font = font_body_sm
            cell.alignment = ac
    ws2.row_dimensions[r].height = 50
    r += 1

# Additional note
r += 1
ws2.merge_cells(f"B{r}:H{r}")
ws2.cell(row=r, column=2).value = "NOTE: Rates vary significantly by region, contract vehicle, and cost accounting structure. Always verify against contractor's most recent FPRA/FPRR. DCAA Forward Pricing Rate Agreements are the gold standard."
ws2.cell(row=r, column=2).font = Font(name="Calibri", size=8, italic=True, bold=True, color=TEAL)
ws2.cell(row=r, column=2).alignment = al
ws2.row_dimensions[r].height = 24

r += 1
ws2.merge_cells(f"B{r}:H{r}")
ws2.cell(row=r, column=2).value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com"
ws2.cell(row=r, column=2).font = font_footer
ws2.cell(row=r, column=2).alignment = ac
ws2.print_area = f"A1:H{r}"

# ═══════════════════════════════════════════════════
# SHEET 3: INSTRUCTIONS
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Instructions")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws3.page_setup.orientation = "portrait"
ws3.page_setup.paperSize = ws3.PAPERSIZE_A4
ws3.page_setup.fitToHeight = 1
ws3.page_setup.fitToWidth = 1

widths3 = {'A': 2, 'B': 4, 'C': 80, 'D': 2}
for c, w in widths3.items():
    ws3.column_dimensions[c].width = w

# Title
r = 1
apply_fill(ws3, r, 1, 4, fill_navy)
ws3.merge_cells("B1:C1")
ws3["B1"].value = "Understanding Wrap Rates in Defense Contracting"
ws3["B1"].font = font_title
ws3["B1"].alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 30

r = 2
apply_fill(ws3, r, 1, 4, fill_navy)
ws3.merge_cells("B2:C2")
ws3["B2"].value = "Acqlerate — Defense Acquisitions Academy"
ws3["B2"].font = font_subtitle
ws3.row_dimensions[2].height = 16

# Content sections
sections = [
    ("WHAT WRAP RATES ARE & WHY THEY MATTER", [
        "A wrap rate (or burden rate/multiplier) is the total cost the government pays per hour of labor divided by the employee's base hourly wage. It captures all indirect costs, overhead, and profit layered on top of direct labor.",
        "",
        "Why it matters to acquisition professionals:",
        "  ▸ It determines the true cost of every labor hour on your contract",
        "  ▸ A $65/hr engineer may cost the government $155-200/hr fully loaded",
        "  ▸ Understanding wrap rates is essential for cost realism analysis in source selection",
        "  ▸ Wrap rate differences between competitors explain most of the price spread in professional services",
        "  ▸ Unrealistically low wrap rates signal an unrealistic proposal (buying in)",
    ]),
    ("HOW DCAA VALIDATES RATES", [
        "The Defense Contract Audit Agency (DCAA) validates contractor indirect rates through several mechanisms:",
        "",
        "  ▸ FPRA (Forward Pricing Rate Agreement) — A written agreement between the contractor and the government establishing billing rates for a future period. Gold standard — means rates are pre-negotiated and DCAA-approved.",
        "  ▸ FPRR (Forward Pricing Rate Recommendation) — When no FPRA exists, DCAA issues a recommendation for rates to use in negotiations. Less authoritative than FPRA.",
        "  ▸ Incurred Cost Audits — Annual audit of actual indirect rates vs. provisional billing rates. Determines final rates and any required adjustments.",
        "  ▸ CAS Compliance — Cost Accounting Standards (CAS 401-420) require consistency in how costs are estimated, accumulated, and reported.",
        "",
        "Key point: If a contractor does NOT have a DCAA-audited rate structure, require them to provide their rate buildup with supporting documentation. Provisional rates may apply.",
    ]),
    ("WHAT 'UNCOMPETITIVE WRAP RATES' MEANS IN SOURCE SELECTION", [
        "During source selection, evaluators performing cost realism analysis should examine wrap rates to ensure:",
        "",
        "  ▸ Rates are consistent with the contractor's FPRA/FPRR (if available)",
        "  ▸ Rates are within industry norms for the contractor's size and type",
        "  ▸ Rates adequately cover the contractor's real cost of doing business",
        "  ▸ Unusually LOW rates may indicate buy-in pricing (contractor plans to raise rates later)",
        "  ▸ Unusually HIGH rates may indicate over-staffing or inefficiency",
        "",
        "Red flags in wrap rate analysis:",
        "  ▸ Wrap rate <1.8x for a mid-size contractor (likely unsustainable)",
        "  ▸ Overhead rate declining year-over-year without explanation",
        "  ▸ Fringe rate below 25% (may indicate inadequate employee benefits / high turnover risk)",
        "  ▸ G&A rate >20% (indicates potential cost structure problems)",
    ]),
    ("HOW TO CHECK IF A COMPETITOR'S PRICING IS OVER-LEVELED", [
        "Over-leveling occurs when a contractor proposes senior/expert labor for work that could be done by junior staff — a common pricing tactic:",
        "",
        "  ▸ Compare proposed labor categories against the PWS/SOW task complexity",
        "  ▸ Check if the proposed labor mix matches industry norms for similar work",
        "  ▸ Verify years of experience claims are realistic for proposed categories",
        "  ▸ Compare against GSA Schedule rates for equivalent labor categories",
        "  ▸ Calculate effective hourly rates and compare against BLS data for the region",
        "",
        "Tip: Request detailed labor category descriptions with minimum qualifications. Map each category to specific PWS tasks. If Level IV engineers are proposed for Level II work, that's over-leveling.",
    ]),
    ("KEY REGULATORY REFERENCES", [
        "  ▸ FAR 31.201 — Composition of total cost (direct + indirect + fee)",
        "  ▸ FAR 31.203 — Indirect costs — principles for allocating indirect costs",
        "  ▸ FAR 31.205 — Selected costs — allowable vs. unallowable cost elements",
        "  ▸ FAR 15.404-1(d) — Cost realism analysis requirements",
        "  ▸ CAS 401 — Consistency in estimating, accumulating, and reporting costs",
        "  ▸ CAS 402 — Consistency in allocating costs incurred for the same purpose",
        "  ▸ CAS 410 — Allocation of business unit G&A expenses",
        "  ▸ CAS 418 — Allocation of direct and indirect costs",
        "  ▸ DFARS 215.404-71 — Weighted Guidelines for profit/fee determination",
        "  ▸ DCAM (DCAA Contract Audit Manual) — Chapter 6: Incurred Costs; Chapter 9: Forward Pricing",
    ]),
]

r = 4
for title, lines in sections:
    ws3.merge_cells(f"B{r}:C{r}")
    ws3.cell(row=r, column=2).value = title
    ws3.cell(row=r, column=2).font = font_section
    ws3.cell(row=r, column=2).fill = fill_teal
    ws3.cell(row=r, column=2).alignment = ac
    apply_fill(ws3, r, 2, 3, fill_teal)
    ws3.row_dimensions[r].height = 22
    r += 1
    
    for line in lines:
        ws3.merge_cells(f"B{r}:C{r}")
        cell = ws3.cell(row=r, column=2)
        if line == "":
            ws3.row_dimensions[r].height = 6
        elif line.startswith("  ▸"):
            cell.value = line
            cell.font = font_inst_body
            cell.alignment = al_top
            ws3.row_dimensions[r].height = 16 if len(line) < 80 else 28
        elif line.endswith(":"):
            cell.value = line
            cell.font = font_inst_bold
            cell.alignment = al
            ws3.row_dimensions[r].height = 16
        else:
            cell.value = line
            cell.font = font_inst_body
            cell.alignment = al_top
            ws3.row_dimensions[r].height = 16 if len(line) < 80 else 28
        r += 1
    r += 1  # Gap between sections

# Footer
ws3.merge_cells(f"B{r}:C{r}")
ws3.cell(row=r, column=2).value = "© 2026 Acqlerate — Defense Acquisitions Academy  |  acqlerate.com  |  For educational use — not legal or financial advice"
ws3.cell(row=r, column=2).font = font_footer
ws3.cell(row=r, column=2).alignment = ac

ws3.print_area = f"A1:D{r}"

# ── Save ──
output = "/home/user/workspace/acq-pro/products/pack3-finance-cheat-sheets/wrap-rate-breakdown.xlsx"
wb.save(output)
print(f"Saved: {output}")
